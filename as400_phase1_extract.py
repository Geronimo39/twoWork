"""
AS/400 Phase 1 – Systemkatalog-Extraktion
==========================================
Zieht alle relevanten Metadaten aus qsys2-Systemkatalogen
und speichert sie als Excel-Datei (ein Sheet pro Quelle).

Voraussetzungen:
    pip install pyodbc pandas openpyxl

Nutzung:
    1. DSN_NAME anpassen (dein ODBC-Datenquellenname)
    2. Optional: SCHEMA_FILTER anpassen (leere Liste = alles)
    3. python as400_phase1_extract.py
"""

import pyodbc
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import sys

# ─────────────────────────────────────────────
# KONFIGURATION – hier anpassen
# ─────────────────────────────────────────────
DSN_NAME      = "DEIN_DSN_NAME"          # <– ODBC-DSN eintragen
USER          = "DEIN_USER"              # <– optional, falls DSN keinen User enthält
PASSWORD      = "DEIN_PASSWORT"          # <– optional

# Leere Liste = alle Schemas. Oder z.B. ["MEINLIB", "PRODLIB"]
SCHEMA_FILTER = []

OUTPUT_FILE   = f"as400_phase1_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
# ─────────────────────────────────────────────


def connect():
    conn_str = f"DSN={DSN_NAME}"
    if USER:
        conn_str += f";UID={USER};PWD={PASSWORD}"
    try:
        conn = pyodbc.connect(conn_str, timeout=30)
        print(f"✓ Verbunden mit: {DSN_NAME}")
        return conn
    except Exception as e:
        print(f"✗ Verbindungsfehler: {e}")
        sys.exit(1)


def schema_where(col="TABLE_SCHEMA"):
    """Erzeugt WHERE-Klausel für Schema-Filter."""
    if not SCHEMA_FILTER:
        return f"{col} NOT IN ('QSYS', 'QSYS2', 'SYSIBM', 'SYSPROC', 'SYSTOOLS')"
    vals = ", ".join(f"'{s}'" for s in SCHEMA_FILTER)
    return f"{col} IN ({vals})"


QUERIES = {

    "00_Übersicht_Schemas": f"""
        SELECT
            TABLE_SCHEMA                        AS schema_name,
            COUNT(*)                            AS anzahl_tabellen,
            SUM(CASE WHEN TABLE_TYPE='T' THEN 1 ELSE 0 END) AS tabellen,
            SUM(CASE WHEN TABLE_TYPE='V' THEN 1 ELSE 0 END) AS views,
            SUM(CASE WHEN TABLE_TYPE='A' THEN 1 ELSE 0 END) AS aliases
        FROM qsys2.systables
        WHERE {schema_where()}
        GROUP BY TABLE_SCHEMA
        ORDER BY anzahl_tabellen DESC
    """,

    "01_Tabellen": f"""
        SELECT
            t.TABLE_SCHEMA                      AS schema_name,
            t.TABLE_NAME                        AS tabelle,
            t.TABLE_TYPE                        AS typ,
            t.TABLE_TEXT                        AS beschreibung,
            t.COLUMN_COUNT                      AS spalten_anzahl,
            COALESCE(s.NUMBER_ROWS, 0)          AS zeilen_anzahl,
            COALESCE(s.DATA_SIZE, 0) / 1024     AS daten_kb,
            t.LAST_ALTERED_TIMESTAMP            AS letzte_aenderung,
            t.SYSTEM_TABLE_NAME                 AS systemname
        FROM qsys2.systables t
        LEFT JOIN qsys2.systablestat s
            ON t.TABLE_SCHEMA = s.TABLE_SCHEMA
            AND t.TABLE_NAME  = s.TABLE_NAME
        WHERE {schema_where('t.TABLE_SCHEMA')}
          AND t.TABLE_TYPE IN ('T', 'P')
        ORDER BY schema_name, tabelle
    """,

    "02_Spalten": f"""
        SELECT
            c.TABLE_SCHEMA                      AS schema_name,
            c.TABLE_NAME                        AS tabelle,
            c.ORDINAL_POSITION                  AS position,
            c.COLUMN_NAME                       AS spalte,
            c.SYSTEM_COLUMN_NAME                AS system_spalte,
            c.DATA_TYPE                         AS datentyp,
            c.LENGTH                            AS laenge,
            c.NUMERIC_SCALE                     AS dezimalstellen,
            c.IS_NULLABLE                       AS nullable,
            c.COLUMN_DEFAULT                    AS standard_wert,
            c.COLUMN_TEXT                       AS beschreibung,
            c.HAS_DEFAULT                       AS hat_default
        FROM qsys2.syscolumns c
        WHERE {schema_where('c.TABLE_SCHEMA')}
        ORDER BY schema_name, tabelle, position
    """,

    "03_Views": f"""
        SELECT
            v.TABLE_SCHEMA                      AS schema_name,
            v.TABLE_NAME                        AS view_name,
            v.VIEW_DEFINITION                   AS view_definition,
            t.TABLE_TEXT                        AS beschreibung,
            t.COLUMN_COUNT                      AS spalten_anzahl
        FROM qsys2.sysviews v
        JOIN qsys2.systables t
            ON v.TABLE_SCHEMA = t.TABLE_SCHEMA
            AND v.TABLE_NAME  = t.TABLE_NAME
        WHERE {schema_where('v.TABLE_SCHEMA')}
        ORDER BY schema_name, view_name
    """,

    "04_Indizes": f"""
        SELECT
            i.TABLE_SCHEMA                      AS schema_name,
            i.TABLE_NAME                        AS tabelle,
            i.INDEX_NAME                        AS index_name,
            i.IS_UNIQUE                         AS eindeutig,
            i.INDEX_TYPE                        AS index_typ,
            i.NUMBER_OF_COLUMNS                 AS anzahl_spalten,
            i.SYSTEM_INDEX_NAME                 AS system_name,
            i.INDEX_TEXT                        AS beschreibung
        FROM qsys2.sysindexes i
        WHERE {schema_where('i.TABLE_SCHEMA')}
        ORDER BY schema_name, tabelle, index_name
    """,

    "05_Schluessel": f"""
        SELECT
            k.TABLE_SCHEMA                      AS schema_name,
            k.TABLE_NAME                        AS tabelle,
            k.INDEX_NAME                        AS index_name,
            k.COLUMN_NAME                       AS spalte,
            k.ORDINAL_POSITION                  AS position,
            k.ORDERING                          AS sortierung
        FROM qsys2.syskeys k
        WHERE {schema_where('k.TABLE_SCHEMA')}
        ORDER BY schema_name, tabelle, index_name, position
    """,

    "06_Tabellenstatistik": f"""
        SELECT
            s.TABLE_SCHEMA                      AS schema_name,
            s.TABLE_NAME                        AS tabelle,
            s.NUMBER_ROWS                       AS zeilen_anzahl,
            s.DATA_SIZE                         AS daten_bytes,
            s.DATA_SIZE / 1024 / 1024           AS daten_mb,
            s.OVERFLOW                          AS overflow_zeilen,
            s.LAST_USED_TIMESTAMP               AS zuletzt_genutzt,
            s.STATISTICS_TIMESTAMP              AS statistik_stand
        FROM qsys2.systablestat s
        WHERE {schema_where('s.TABLE_SCHEMA')}
          AND s.NUMBER_ROWS > 0
        ORDER BY s.NUMBER_ROWS DESC
    """,

    "07_Spaltenstatistik": f"""
        SELECT
            cs.TABLE_SCHEMA                     AS schema_name,
            cs.TABLE_NAME                       AS tabelle,
            cs.COLUMN_NAME                      AS spalte,
            cs.DISTINCT_COUNT                   AS eindeutige_werte,
            cs.NULL_COUNT                       AS null_anzahl,
            cs.STATISTICS_TIMESTAMP             AS statistik_stand
        FROM qsys2.syscolumnstat cs
        WHERE {schema_where('cs.TABLE_SCHEMA')}
        ORDER BY schema_name, tabelle, spalte
    """,

    "08_Programme": f"""
        SELECT
            p.ROUTINE_SCHEMA                    AS schema_name,
            p.ROUTINE_NAME                      AS programm,
            p.ROUTINE_TYPE                      AS typ,
            p.EXTERNAL_NAME                     AS externer_name,
            p.LANGUAGE                          AS sprache,
            p.ROUTINE_TEXT                      AS beschreibung,
            p.LAST_ALTERED_TIMESTAMP            AS letzte_aenderung,
            p.CREATED                           AS erstellt
        FROM qsys2.sysroutines p
        WHERE {schema_where('p.ROUTINE_SCHEMA')}
        ORDER BY schema_name, programm
    """,
}


def run_query(conn, name, sql):
    print(f"  → Abfrage: {name} ...", end=" ", flush=True)
    try:
        df = pd.read_sql(sql, conn)
        print(f"✓ {len(df):,} Zeilen")
        return df
    except Exception as e:
        print(f"⚠ Fehler: {e}")
        return pd.DataFrame({"Fehler": [str(e)]})


def style_sheet(ws, df, tab_color):
    """Formatiert ein Worksheet professionell."""
    header_fill   = PatternFill("solid", fgColor="1F4E79")
    header_font   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    data_font     = Font(name="Arial", size=10)
    alt_fill      = PatternFill("solid", fgColor="EBF3FB")
    center_align  = Alignment(horizontal="center", vertical="center", wrap_text=False)
    left_align    = Alignment(horizontal="left",   vertical="center", wrap_text=False)
    thin          = Side(style="thin", color="D0D0D0")
    border        = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.sheet_properties.tabColor = tab_color
    ws.freeze_panes = "A2"

    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name.upper())
        cell.font    = header_font
        cell.fill    = header_fill
        cell.border  = border
        cell.alignment = center_align

    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        fill = alt_fill if row_idx % 2 == 0 else None
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font   = data_font
            cell.border = border
            cell.alignment = left_align
            if fill:
                cell.fill = fill

    for col_idx, col_name in enumerate(df.columns, 1):
        max_len = max(
            len(str(col_name)),
            df[col_name].astype(str).str.len().max() if len(df) else 0
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)

    ws.row_dimensions[1].height = 20


TAB_COLORS = ["2E75B6", "00B050", "7030A0", "FF8C00",
              "C00000", "4BACC6", "F79646", "0070C0", "70AD47"]


def main():
    print("\n=== AS/400 Phase 1 – Systemkatalog-Extraktion ===\n")

    conn = connect()

    results = {}
    for name, sql in QUERIES.items():
        results[name] = run_query(conn, name, sql)

    conn.close()
    print("\n✓ Verbindung geschlossen")

    print(f"\nSchreibe Excel: {OUTPUT_FILE} ...")
    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        for name, df in results.items():
            df.to_excel(writer, sheet_name=name[:31], index=False)

    wb = load_workbook(OUTPUT_FILE)
    for i, ws_name in enumerate(wb.sheetnames):
        ws  = wb[ws_name]
        df  = results.get(ws_name, pd.DataFrame())
        style_sheet(ws, df, TAB_COLORS[i % len(TAB_COLORS)])

    # Summary-Sheet
    ws_sum = wb.create_sheet("00_Summary", 0)
    ws_sum.sheet_properties.tabColor = "1F4E79"
    ws_sum["A1"] = "AS/400 Systemkatalog-Analyse – Phase 1"
    ws_sum["A1"].font = Font(name="Arial", bold=True, size=14, color="1F4E79")
    ws_sum["A3"] = "Extraktionsdatum:"
    ws_sum["B3"] = datetime.now().strftime("%d.%m.%Y %H:%M")
    ws_sum["A4"] = "DSN:"
    ws_sum["B4"] = DSN_NAME

    headers = ["Sheet", "Beschreibung", "Datensätze"]
    descriptions = {
        "00_Übersicht_Schemas": "Schemas / Libraries im Überblick",
        "01_Tabellen":          "Alle physischen Tabellen mit Größe",
        "02_Spalten":           "Alle Spalten aller Tabellen",
        "03_Views":             "Views inkl. Definition",
        "04_Indizes":           "Alle Indizes",
        "05_Schluessel":        "Index-Schlüsselspalten",
        "06_Tabellenstatistik": "Zeilenzahl & Größe (top-down sortiert)",
        "07_Spaltenstatistik":  "Distinct-Werte & NULL-Anteile",
        "08_Programme":         "Programme / Routinen",
    }
    ws_sum["A6"] = "SHEET"
    ws_sum["B6"] = "BESCHREIBUNG"
    ws_sum["C6"] = "DATENSÄTZE"
    for col in ["A6","B6","C6"]:
        ws_sum[col].font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        ws_sum[col].fill = PatternFill("solid", fgColor="1F4E79")

    for r, (sheet, desc) in enumerate(descriptions.items(), 7):
        ws_sum.cell(r, 1, sheet).font  = Font(name="Arial", size=10)
        ws_sum.cell(r, 2, desc).font   = Font(name="Arial", size=10)
        cnt = len(results.get(sheet, pd.DataFrame()))
        ws_sum.cell(r, 3, cnt).font    = Font(name="Arial", size=10)
        ws_sum.cell(r, 3).alignment    = Alignment(horizontal="right")

    ws_sum.column_dimensions["A"].width = 28
    ws_sum.column_dimensions["B"].width = 42
    ws_sum.column_dimensions["C"].width = 14

    wb.save(OUTPUT_FILE)
    print(f"\n✓ Fertig! Datei gespeichert: {OUTPUT_FILE}")
    print("\nNächster Schritt: as400_phase1_analyse.ipynb öffnen\n")


if __name__ == "__main__":
    main()
